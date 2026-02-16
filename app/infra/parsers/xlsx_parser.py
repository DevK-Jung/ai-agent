"""
XLSX 파서 구현
"""
from typing import List
from .base import BaseParser
from .models import ParsedContent, Table, Image


class XLSXParser(BaseParser):
    """XLSX 파일 전용 파서"""
    
    def get_supported_mime_types(self) -> List[str]:
        return [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ]
    
    async def parse(self, file_path: str) -> ParsedContent:
        """XLSX 파일에서 구조화된 콘텐츠를 추출합니다."""
        if not self.validate_file(file_path):
            raise ValueError(f"유효하지 않은 XLSX 파일: {file_path}")
        
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            full_text = ""
            tables = []
            structure = {"sheets": [], "total_rows": 0, "total_cols": 0}
            
            # 모든 시트 처리
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                
                # 시트의 실제 사용 범위 확인
                if sheet.max_row == 1 and sheet.max_column == 1:
                    # 빈 시트인 경우 건너뛰기
                    if not sheet.cell(1, 1).value:
                        continue
                
                # 시트 정보를 구조에 추가
                sheet_info = {
                    "name": sheet_name,
                    "index": sheet_idx,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column
                }
                structure["sheets"].append(sheet_info)
                structure["total_rows"] += sheet.max_row
                structure["total_cols"] = max(structure["total_cols"], sheet.max_column)
                
                # 시트 데이터를 테이블로 추출
                sheet_data = []
                headers = []
                
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    # None 값들을 빈 문자열로 변환
                    clean_row = [str(cell) if cell is not None else "" for cell in row]
                    
                    # 모든 셀이 비어있는 행은 건너뛰기
                    if not any(clean_row):
                        continue
                    
                    if row_idx == 1:
                        # 첫 번째 행을 헤더로 사용
                        headers = clean_row
                        # 헤더가 모두 비어있으면 컬럼명 생성
                        if not any(headers):
                            headers = [get_column_letter(i) for i in range(1, len(clean_row) + 1)]
                    else:
                        sheet_data.append(clean_row)
                
                # 테이블 생성 (데이터가 있는 경우만)
                if headers or sheet_data:
                    tables.append(Table(
                        headers=headers,
                        rows=sheet_data,
                        position={"sheet": sheet_name, "sheet_index": sheet_idx},
                        metadata={
                            "extraction_method": "openpyxl",
                            "sheet_name": sheet_name,
                            "max_row": sheet.max_row,
                            "max_column": sheet.max_column,
                            "data_rows": len(sheet_data)
                        }
                    ))
                
                # 텍스트 형태로도 추가 (검색/청킹용)
                full_text += f"[시트: {sheet_name}]\n"
                
                # 헤더 추가
                if headers:
                    full_text += " | ".join(headers) + "\n"
                    full_text += "-" * len(" | ".join(headers)) + "\n"
                
                # 데이터 행 추가 (처음 몇 행만, 너무 길어지지 않도록)
                max_preview_rows = 20
                for row_data in sheet_data[:max_preview_rows]:
                    # 빈 행 건너뛰기
                    if any(cell.strip() for cell in row_data if isinstance(cell, str)):
                        full_text += " | ".join(row_data) + "\n"
                
                if len(sheet_data) > max_preview_rows:
                    full_text += f"... ({len(sheet_data) - max_preview_rows}행 더 있음)\n"
                
                full_text += "\n"
            
            # 이미지/차트 정보 추출 (기본적인 정보만)
            images = []
            try:
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # 이미지 추출
                    if hasattr(sheet, '_images') and sheet._images:
                        for img_idx, img in enumerate(sheet._images):
                            images.append(Image(
                                name=f"{sheet_name}_image_{img_idx + 1}",
                                format="unknown",  # openpyxl에서 정확한 format 추출은 복잡
                                position={"sheet": sheet_name},
                                metadata={"extraction_method": "openpyxl"}
                            ))
                    
                    # 차트 정보
                    if hasattr(sheet, '_charts') and sheet._charts:
                        for chart_idx, chart in enumerate(sheet._charts):
                            images.append(Image(
                                name=f"{sheet_name}_chart_{chart_idx + 1}",
                                format="chart",
                                position={"sheet": sheet_name},
                                metadata={
                                    "extraction_method": "openpyxl",
                                    "type": "chart"
                                }
                            ))
            except Exception as e:
                self.logger.warning(f"XLSX 이미지/차트 정보 추출 실패: {e}")
            
            # 메타데이터 추출
            metadata = {
                "parser": "openpyxl",
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "total_tables": len(tables)
            }
            
            # 문서 속성 추출
            try:
                props = workbook.properties
                metadata.update({
                    "title": props.title or "",
                    "author": props.creator or "",
                    "subject": props.subject or "",
                    "created": str(props.created) if props.created else "",
                    "modified": str(props.modified) if props.modified else ""
                })
            except Exception as e:
                self.logger.warning(f"XLSX 메타데이터 추출 실패: {e}")
            
            workbook.close()
            
            return ParsedContent(
                raw_text=full_text.strip(),
                metadata=metadata,
                structure=structure,
                tables=tables,
                images=images
            )
            
        except ImportError:
            raise ImportError("openpyxl이 설치되지 않았습니다: pip install openpyxl")
        except Exception as e:
            raise ValueError(f"XLSX 파싱 실패: {e}")